from django.shortcuts import render, redirect
from .models import *
from .forms import *
import subprocess
import openpyxl

# Create your views here.
def home(request):
    drives = RepoStatus.objects.all()
    form = DriveForm()
    print(drives)
    if request.method == 'POST':
        form = DriveForm(request.POST)
        if form.is_valid():
            form.save()
        return redirect('/')

    context = { 'drives': drives, 'form': form }
    return render(request, 'diskUsage/home.html', context)
    # return render(request, 'diskUsage/home2.html', context)

def create(request):
    form = DriveForm()
    context = { 'form': form }
    return render(request, 'diskUsage/create.html', context)

def service(request, id=0):
    if id == '1':
        pspath = r'C:\Windows\System32\WindowsPowerShell\v1.0\powershell.exe'
        subprocess.getoutput(pspath + ' Start-Service -Name PythonDU')
        output = subprocess.getoutput(pspath + ' Get-Service PythonDU').split()[6]
        mylist = []
        mydict = {}
        mydict['status'] = output
        mylist.append(mydict)
        context = {'mylist': mylist}
        return render(request, 'diskUsage/service.html', context)

    elif id == '2':
        pspath = r'C:\Windows\System32\WindowsPowerShell\v1.0\powershell.exe'
        subprocess.getoutput(pspath + ' Stop-Service -Name PythonDU')
        output = subprocess.getoutput(pspath + ' Get-Service PythonDU').split()[6]
        mylist = []
        mydict = {}
        mydict['status'] = output
        mylist.append(mydict)
        context = {'mylist': mylist}
        return render(request, 'diskUsage/service.html', context)

    else:
        pspath = r'C:\Windows\System32\WindowsPowerShell\v1.0\powershell.exe'
        output = subprocess.getoutput(pspath + ' Get-Service PythonDU').split()[6]
        mylist = []
        mydict = {}
        mydict['status'] = output
        mylist.append(mydict)
        context = {'mylist': mylist}
        return render(request, 'diskUsage/service.html', context)

def upload(request):
    if "GET" == request.method:
        return render(request, 'diskUsage/upload.html', {})
    else:
        excel_file = request.FILES["excel_file"]

        # you may put validations here to check extension or file size

        # loading excel file
        wb = openpyxl.load_workbook(excel_file)

        # getting sheets names
        sheets = wb.sheetnames

        # getting a particular sheet
        # worksheet = wb["Sheet1"]
        worksheet = wb[sheets[0]]

        # getting active sheet
        # active_sheet = wb.active
        # print(active_sheet)

        # reading a cell
        # print(worksheet["A1"].value)

        excel_data = list()
        # iterating over the rows and getting value from each cell in row
        for row in worksheet.iter_rows():
            row_data = list()

            for cell in row:
                row_data.append(str(cell.value))

            excel_data.append(row_data)


        # removing header from excel dataset
        exceluploadData = []
        for row in excel_data:
            newdict = {}
            newdict = {}
            newdict['path'] = row[0]
            exceluploadData.append(newdict)

        cleaned_excel_data = []

        for val in excel_data:
            if val[0].lower() not in ['none', 'path']:
                cleaned_excel_data.append(val)
            else:
                continue

        excel_data = cleaned_excel_data

        # Posting Data to database
        for row in excel_data:
            _, created = CurrentJobs.objects.update_or_create(
            path = row[0])

        context = { "excel_data": excel_data, 'exceluploadData': exceluploadData }

        return render(request, 'diskUsage/upload.html', context)
